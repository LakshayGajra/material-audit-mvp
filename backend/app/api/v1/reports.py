"""
Reports API Endpoints

Provides Excel downloads for various reports.
"""
import io
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import (
    Anomaly,
    Audit,
    AuditLineItem,
    Contractor,
    ContractorInventory,
    Consumption,
    InventoryAdjustment,
    Material,
    MaterialIssuance,
    MaterialRejection,
    Reconciliation,
    ReconciliationLine,
    Warehouse,
    WarehouseInventory,
)

router = APIRouter(prefix="/api/v1/reports", tags=["Reports"])


def create_excel_response(workbook, filename: str):
    """Create a streaming response for an Excel file."""
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/inventory-summary")
def get_inventory_summary_report(db: Session = Depends(get_db)):
    """
    Download inventory summary across all warehouses and contractors.
    Returns an Excel file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed. Please install it to generate Excel reports."
        )

    wb = Workbook()

    # Sheet 1: Warehouse Inventory
    ws1 = wb.active
    ws1.title = "Warehouse Inventory"
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    headers = ["Warehouse", "Material Code", "Material Name", "Current Qty", "Unit", "Reorder Point", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    warehouse_inventory = db.query(WarehouseInventory).all()
    for row_num, inv in enumerate(warehouse_inventory, 2):
        warehouse = db.query(Warehouse).filter(Warehouse.id == inv.warehouse_id).first()
        material = db.query(Material).filter(Material.id == inv.material_id).first()
        status = "Low Stock" if inv.current_quantity <= inv.reorder_point else "OK"

        ws1.cell(row=row_num, column=1, value=warehouse.name if warehouse else "Unknown")
        ws1.cell(row=row_num, column=2, value=material.code if material else "Unknown")
        ws1.cell(row=row_num, column=3, value=material.name if material else "Unknown")
        ws1.cell(row=row_num, column=4, value=float(inv.current_quantity) if inv.current_quantity else 0)
        ws1.cell(row=row_num, column=5, value=inv.unit_of_measure)
        ws1.cell(row=row_num, column=6, value=float(inv.reorder_point) if inv.reorder_point else 0)
        ws1.cell(row=row_num, column=7, value=status)

    # Sheet 2: Contractor Inventory
    ws2 = wb.create_sheet("Contractor Inventory")
    headers = ["Contractor", "Material Code", "Material Name", "Quantity"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    contractor_inventory = db.query(ContractorInventory).all()
    for row_num, inv in enumerate(contractor_inventory, 2):
        contractor = db.query(Contractor).filter(Contractor.id == inv.contractor_id).first()
        material = db.query(Material).filter(Material.id == inv.material_id).first()

        ws2.cell(row=row_num, column=1, value=contractor.name if contractor else "Unknown")
        ws2.cell(row=row_num, column=2, value=material.code if material else "Unknown")
        ws2.cell(row=row_num, column=3, value=material.name if material else "Unknown")
        ws2.cell(row=row_num, column=4, value=float(inv.quantity) if inv.quantity else 0)

    filename = f"inventory_summary_{date.today().isoformat()}.xlsx"
    return create_excel_response(wb, filename)


@router.get("/material-movement/{material_id}")
def get_material_movement_report(
    material_id: int,
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
):
    """
    Track all movements of a specific material.
    Returns a timeline of issuances, consumption, and rejections.
    """
    material = db.query(Material).filter(Material.id == material_id).first()
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")

    movements = []

    # Issuances
    issuance_query = db.query(MaterialIssuance).filter(MaterialIssuance.material_id == material_id)
    if date_from:
        issuance_query = issuance_query.filter(MaterialIssuance.issued_date >= date_from)
    if date_to:
        issuance_query = issuance_query.filter(MaterialIssuance.issued_date <= date_to)

    for iss in issuance_query.all():
        contractor = db.query(Contractor).filter(Contractor.id == iss.contractor_id).first()
        movements.append({
            "date": str(iss.issued_date),
            "type": "ISSUANCE",
            "quantity": float(iss.quantity),
            "direction": "OUT",
            "entity": contractor.name if contractor else "Unknown",
            "reference": iss.issuance_number,
            "notes": f"Issued by {iss.issued_by}",
        })

    # Consumption
    consumption_query = db.query(Consumption).filter(Consumption.material_id == material_id)
    for cons in consumption_query.all():
        contractor = db.query(Contractor).filter(Contractor.id == cons.contractor_id).first()
        movements.append({
            "date": str(date.today()),  # Consumption doesn't have date field directly
            "type": "CONSUMPTION",
            "quantity": float(cons.quantity),
            "direction": "CONSUMED",
            "entity": contractor.name if contractor else "Unknown",
            "reference": f"Production #{cons.production_record_id}",
            "notes": "Used in production",
        })

    # Rejections
    rejection_query = db.query(MaterialRejection).filter(MaterialRejection.material_id == material_id)
    if date_from:
        rejection_query = rejection_query.filter(MaterialRejection.rejection_date >= date_from)
    if date_to:
        rejection_query = rejection_query.filter(MaterialRejection.rejection_date <= date_to)

    for rej in rejection_query.all():
        contractor = db.query(Contractor).filter(Contractor.id == rej.contractor_id).first()
        movements.append({
            "date": str(rej.rejection_date),
            "type": "REJECTION",
            "quantity": float(rej.quantity),
            "direction": "RETURN",
            "entity": contractor.name if contractor else "Unknown",
            "reference": rej.rejection_number,
            "notes": rej.rejection_reason,
        })

    # Sort by date
    movements.sort(key=lambda x: x["date"], reverse=True)

    return {
        "material": {
            "id": material.id,
            "code": material.code,
            "name": material.name,
            "unit": material.unit,
        },
        "date_range": {
            "from": str(date_from) if date_from else None,
            "to": str(date_to) if date_to else None,
        },
        "movements": movements,
        "summary": {
            "total_issued": sum(m["quantity"] for m in movements if m["type"] == "ISSUANCE"),
            "total_consumed": sum(m["quantity"] for m in movements if m["type"] == "CONSUMPTION"),
            "total_rejected": sum(m["quantity"] for m in movements if m["type"] == "REJECTION"),
        },
    }


@router.get("/contractor-audit-history/{contractor_id}")
def get_contractor_audit_history(
    contractor_id: int,
    db: Session = Depends(get_db),
):
    """
    Download audit and reconciliation history for a contractor.
    Returns an Excel file.
    """
    contractor = db.query(Contractor).filter(Contractor.id == contractor_id).first()
    if not contractor:
        raise HTTPException(status_code=404, detail="Contractor not found")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed"
        )

    wb = Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Sheet 1: Audits
    ws1 = wb.active
    ws1.title = "Audits"
    headers = ["Audit #", "Date", "Auditor", "Status", "Material", "Expected", "Physical", "Variance", "Variance %"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    audits = db.query(Audit).filter(Audit.contractor_id == contractor_id).all()
    row_num = 2
    for audit in audits:
        line_items = db.query(AuditLineItem).filter(AuditLineItem.audit_id == audit.id).all()
        for item in line_items:
            material = db.query(Material).filter(Material.id == item.material_id).first()
            ws1.cell(row=row_num, column=1, value=audit.audit_number)
            ws1.cell(row=row_num, column=2, value=str(audit.audit_date))
            ws1.cell(row=row_num, column=3, value=audit.auditor_name)
            ws1.cell(row=row_num, column=4, value=audit.status)
            ws1.cell(row=row_num, column=5, value=material.name if material else "Unknown")
            ws1.cell(row=row_num, column=6, value=float(item.expected_quantity) if item.expected_quantity else 0)
            ws1.cell(row=row_num, column=7, value=float(item.physical_quantity) if item.physical_quantity else 0)
            ws1.cell(row=row_num, column=8, value=float(item.variance) if item.variance else 0)
            ws1.cell(row=row_num, column=9, value=float(item.variance_percentage) if item.variance_percentage else 0)
            row_num += 1

    # Sheet 2: Reconciliations
    ws2 = wb.create_sheet("Reconciliations")
    headers = ["Recon #", "Date", "Period", "Status", "Material", "System Qty", "Reported Qty", "Variance", "Variance %"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    reconciliations = db.query(Reconciliation).filter(Reconciliation.contractor_id == contractor_id).all()
    row_num = 2
    for recon in reconciliations:
        line_items = db.query(ReconciliationLine).filter(ReconciliationLine.reconciliation_id == recon.id).all()
        for item in line_items:
            material = db.query(Material).filter(Material.id == item.material_id).first()
            ws2.cell(row=row_num, column=1, value=recon.reconciliation_number)
            ws2.cell(row=row_num, column=2, value=str(recon.reconciliation_date))
            ws2.cell(row=row_num, column=3, value=recon.period_type)
            ws2.cell(row=row_num, column=4, value=recon.status)
            ws2.cell(row=row_num, column=5, value=material.name if material else "Unknown")
            ws2.cell(row=row_num, column=6, value=float(item.system_quantity) if item.system_quantity else 0)
            ws2.cell(row=row_num, column=7, value=float(item.reported_quantity) if item.reported_quantity else 0)
            ws2.cell(row=row_num, column=8, value=float(item.variance) if item.variance else 0)
            ws2.cell(row=row_num, column=9, value=float(item.variance_percentage) if item.variance_percentage else 0)
            row_num += 1

    # Sheet 3: Adjustments
    ws3 = wb.create_sheet("Inventory Adjustments")
    headers = ["Adj #", "Date", "Material", "Type", "Qty Before", "Qty After", "Adjustment", "Reason", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    adjustments = db.query(InventoryAdjustment).filter(InventoryAdjustment.contractor_id == contractor_id).all()
    for row_num, adj in enumerate(adjustments, 2):
        material = db.query(Material).filter(Material.id == adj.material_id).first()
        ws3.cell(row=row_num, column=1, value=adj.adjustment_number)
        ws3.cell(row=row_num, column=2, value=str(adj.adjustment_date))
        ws3.cell(row=row_num, column=3, value=material.name if material else "Unknown")
        ws3.cell(row=row_num, column=4, value=adj.adjustment_type)
        ws3.cell(row=row_num, column=5, value=float(adj.quantity_before) if adj.quantity_before else 0)
        ws3.cell(row=row_num, column=6, value=float(adj.quantity_after) if adj.quantity_after else 0)
        ws3.cell(row=row_num, column=7, value=float(adj.adjustment_quantity) if adj.adjustment_quantity else 0)
        ws3.cell(row=row_num, column=8, value=adj.reason)
        ws3.cell(row=row_num, column=9, value=adj.status)

    filename = f"contractor_{contractor.code}_audit_history_{date.today().isoformat()}.xlsx"
    return create_excel_response(wb, filename)


@router.get("/anomaly-report")
def get_anomaly_report(
    status: Optional[str] = Query(None, description="resolved or unresolved"),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
):
    """
    Download anomaly report with filters.
    Returns an Excel file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    query = db.query(Anomaly)

    if status == "resolved":
        query = query.filter(Anomaly.resolved == True)
    elif status == "unresolved":
        query = query.filter(Anomaly.resolved == False)

    if date_from:
        query = query.filter(Anomaly.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        query = query.filter(Anomaly.created_at <= datetime.combine(date_to, datetime.max.time()))

    anomalies = query.order_by(Anomaly.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    headers = [
        "ID", "Contractor", "Material", "Type", "Expected", "Actual",
        "Variance", "Variance %", "Severity", "Status", "Created", "Resolved At", "Notes"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_num, anom in enumerate(anomalies, 2):
        contractor = db.query(Contractor).filter(Contractor.id == anom.contractor_id).first()
        material = db.query(Material).filter(Material.id == anom.material_id).first()

        # Determine severity based on variance percentage
        variance_pct = abs(anom.variance_percent) if anom.variance_percent else 0
        if variance_pct > 20:
            severity = "CRITICAL"
        elif variance_pct > 10:
            severity = "HIGH"
        elif variance_pct > 5:
            severity = "MEDIUM"
        else:
            severity = "LOW"

        ws.cell(row=row_num, column=1, value=anom.id)
        ws.cell(row=row_num, column=2, value=contractor.name if contractor else "Unknown")
        ws.cell(row=row_num, column=3, value=material.name if material else "Unknown")
        ws.cell(row=row_num, column=4, value=anom.anomaly_type)
        ws.cell(row=row_num, column=5, value=anom.expected_quantity)
        ws.cell(row=row_num, column=6, value=anom.actual_quantity)
        ws.cell(row=row_num, column=7, value=anom.variance)
        ws.cell(row=row_num, column=8, value=anom.variance_percent)
        ws.cell(row=row_num, column=9, value=severity)
        ws.cell(row=row_num, column=10, value="Resolved" if anom.resolved else "Open")
        ws.cell(row=row_num, column=11, value=str(anom.created_at) if anom.created_at else "")
        ws.cell(row=row_num, column=12, value=str(anom.resolved_at) if anom.resolved_at else "")
        ws.cell(row=row_num, column=13, value=anom.notes)

    filename = f"anomaly_report_{date.today().isoformat()}.xlsx"
    return create_excel_response(wb, filename)
